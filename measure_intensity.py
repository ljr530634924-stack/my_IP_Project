import csv
import cv2
import numpy as np
import os


def compute_quadrant_intensity(
    brightness_image,
    labels,
    regions,
    axes_info,
    csv_path="quadrant_intensity.csv",
    id_map_path="particle_id_map.png",
):
    """
    Per particle: place one sampling circle in each quadrant (radius = minor_axis_length / 2,
    center offset along ex/ey by the same radius), compute mean intensity inside each circle,
    and draw IDs + circles onto id_map for visual validation.
    """
    h, w = brightness_image.shape

    # 1) stable IDs: sort by x (col) from left to right
    region_sorted = sorted(regions, key=lambda r: r.centroid[1])
    region_id_map = {r.label: idx + 1 for idx, r in enumerate(region_sorted)}

    # 2) base id map
    # Create a displayable 8-bit background for the ID map by normalizing the input brightness image.
    # This handles both 8-bit and 16-bit inputs correctly for visualization purposes,
    # without affecting the original `brightness_image` used for measurement.
    min_val, max_val = brightness_image.min(), brightness_image.max()
    if max_val > min_val:
        # Normalize to 0-255 range
        display_base = 255.0 * (brightness_image - min_val) / (max_val - min_val)
    else:
        # Handle flat image
        display_base = np.zeros_like(brightness_image)
    id_map = cv2.cvtColor(display_base.astype(np.uint8), cv2.COLOR_GRAY2BGR)

    for r in region_sorted:
        pid = region_id_map[r.label]
        cx, cy = r.centroid[1], r.centroid[0]
        cv2.putText(
            id_map,
            str(pid),
            (int(cx), int(cy)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (0, 0, 255),
            2,
        )

    # 3) circles per quadrant
    #Q1 is blue, Q2 is green, Q3 is purple, Q4 is yellow
    results = []
    color_map = {"Q1": (255, 0, 0), "Q2": (0, 255, 0), "Q3": (255, 0, 255), "Q4": (0, 255, 255)}
    offsets = {
        "Q1": (+1, +1),
        "Q2": (-1, +1),
        "Q3": (-1, -1),
        "Q4": (+1, -1),
    }

    for r in region_sorted:
        label = r.label
        if label not in axes_info:
            continue
        pid = region_id_map[label]
        mask = labels == label

        ax = axes_info[label]  # (x, y)
        C_xy = np.array(ax["centroid"], dtype=np.float32)
        ex_xy = np.array(ax["ex"], dtype=np.float32)
        ey_xy = np.array(ax["ey"], dtype=np.float32)

        # convert to (row, col)
        C = np.array([C_xy[1], C_xy[0]], dtype=np.float32)
        ex_rc = np.array([ex_xy[1], ex_xy[0]], dtype=np.float32)
        ey_rc = np.array([ey_xy[1], ey_xy[0]], dtype=np.float32)

        def _norm(v):
            n = np.linalg.norm(v)
            return v if n == 0 else v / n

        ex_rc = _norm(ex_rc)
        ey_rc = _norm(ey_rc)

        radius = float(r.minor_axis_length) / 6
        if radius <= 0:
            results.append([pid, np.nan, np.nan, np.nan, np.nan])
            continue

        ys, xs = np.where(mask)
        coords = np.stack([ys, xs], axis=1).astype(np.float32)
        vals = brightness_image[mask]

        quadrant_means = []
        quadrant_areas = []

        for q_name, (sx, sy) in offsets.items():
            center = C + ex_rc * (sx * radius) + ey_rc * (sy * radius)

            dy = coords[:, 0] - center[0]
            dx = coords[:, 1] - center[1]
            in_circle = (dx * dx + dy * dy) <= (radius * radius)

            q_vals = vals[in_circle]
            if q_vals.size > 0:
                mean_val = round(float(np.mean(q_vals)), 2)
                area_val = q_vals.size
            else:
                mean_val = np.nan
                area_val = 0

            quadrant_means.append(mean_val)
            quadrant_areas.append(area_val)

            cv2.circle(
                id_map,
                (int(round(center[1])), int(round(center[0]))),  # (x, y)
                int(round(radius)),
                color_map[q_name],
                1,
                cv2.LINE_AA,
            )

        results.append([pid, *quadrant_means, *quadrant_areas])

    # 4) write CSV
    mean_row = []
    sd_row = []
    sem_row = []
    if results:
        mean_row = ["Mean"]
        sd_row = ["SD"]
        sem_row = ["Std. Error"]
        # Columns 1 to 8 (Q1_Mean...Q4_Mean, Q1_Area...Q4_Area)
        for col_idx in range(1, 9):
            col_vals = [row[col_idx] for row in results]
            valid_vals = [v for v in col_vals if not np.isnan(v)]
            if valid_vals:
                mean_row.append(round(np.mean(valid_vals), 2))
                n = len(valid_vals)
                if n > 1:
                    sd = np.std(valid_vals, ddof=1)
                    sd_row.append(round(sd, 2))
                    sem_row.append(round(sd / np.sqrt(n), 2))
                else:
                    sd_row.append(0.0)
                    sem_row.append(0.0)
            else:
                mean_row.append(0.0)
                sd_row.append(0.0)
                sem_row.append(0.0)

    # Prepare footers (Mean / SD / SEM rows)
    footer_rows = []
    if mean_row:
        footer_rows.append(mean_row)
    if sd_row:
        footer_rows.append(sd_row)
    if sem_row:
        footer_rows.append(sem_row)

    header = ["particle_id", 
              "Q1_Mean", "Q2_Mean", "Q3_Mean", "Q4_Mean", 
              "Q1_Area", "Q2_Area", "Q3_Area", "Q4_Area"]
    
    # Write colored Excel
    _write_colored_xlsx(csv_path, header, results, footer_rows=footer_rows)

    # 5) Save population summary (Inter-particle stats)
    base, ext = os.path.splitext(csv_path)
    # _save_population_stats(results, f"{base}_summary{ext}")

    cv2.imwrite(id_map_path, id_map)

    print(f"[OK] quadrant table written: {csv_path}")
    print(f"[OK] id map saved: {id_map_path}")


def compute_masked_quadrant_intensity(
    brightness_image,
    structure_mask,
    signal_mask,
    axes_info,
    csv_path="masked_quadrant_results.csv",
    overlay_path="final_overlay_visualization.png"
):
    """
    New simplified logic:
    1. Calculate total signal area S inside the particle (intersection of structure & signal_mask).
    2. Define measurement area per quadrant = S / 4.
    3. Calculate radius r = sqrt((S/4)/pi).
    4. Place 4 circles of radius r at offsets (+/-r, +/-r) from centroid in the local axes system.
    5. Measure mean intensity within these circles.
    """
    h, w = brightness_image.shape

    # 1. Prepare structure labels
    num_labels, labels = cv2.connectedComponents(structure_mask)
    from skimage import measure
    regions = measure.regionprops(labels)
    regions = sorted(regions, key=lambda r: r.centroid[1])
    region_id_map = {r.label: idx + 1 for idx, r in enumerate(regions)}

    # 2. Prepare visualization base (8-bit)
    min_val, max_val = brightness_image.min(), brightness_image.max()
    if max_val > min_val:
        display_base = 255.0 * (brightness_image - min_val) / (max_val - min_val)
    else:
        display_base = np.zeros_like(brightness_image)
    vis_img = cv2.cvtColor(display_base.astype(np.uint8), cv2.COLOR_GRAY2BGR)

    # Quadrant colors for visualization
    # Q1: Blue, Q2: Green, Q3: Purple, Q4: Yellow
    # Format: (Name, sign_x, sign_y, Color)
    quad_defs = [
        ("Q1", 1, 1, (255, 0, 0)),
        ("Q2", -1, 1, (0, 255, 0)),
        ("Q3", -1, -1, (255, 0, 255)),
        ("Q4", 1, -1, (0, 255, 255))
    ]

    results = []

    for r in regions:
        label = r.label
        if label not in axes_info:
            continue

        pid = region_id_map[label]
        ax = axes_info[label]
        centroid = np.array(ax["centroid"])
        ex = np.array(ax["ex"])
        ey = np.array(ax["ey"])

        # --- Step 1: Calculate S (Total Signal Area inside Particle) ---
        minr, minc, maxr, maxc = r.bbox
        
        # Extract local masks to speed up
        minr, minc = max(0, minr), max(0, minc)
        maxr, maxc = min(h, maxr), min(w, maxc)
        
        local_labels = labels[minr:maxr, minc:maxc]
        local_signal = signal_mask[minr:maxr, minc:maxc]
        
        # Intersection: pixel is in this particle AND is signal
        intersection = (local_labels == label) & (local_signal > 0)
        S = np.count_nonzero(intersection)

        means = []
        areas = []

        if S > 0:
            # --- Step 2 & 3: Calculate radius r ---
            # Area per quadrant = S / 4
            # pi * r^2 = S / 4
            target_area = S / 4.0
            radius = np.sqrt(target_area / np.pi)
            
            # --- Step 4: Measure in 4 circles ---
            # Centers at centroid + sx*r*ex + sy*r*ey (tangent to axes logic)
            for q_name, sx, sy, color in quad_defs:
                center_offset = (ex * sx * radius) + (ey * sy * radius)
                center_q = centroid + center_offset
                
                cx_int, cy_int = int(round(center_q[0])), int(round(center_q[1]))
                r_int = int(round(radius))
                
                if r_int <= 0:
                    means.append(0.0)
                    areas.append(0)
                    continue

                # Create mask for this circle
                # Define ROI to avoid scanning full image
                y_min, y_max = max(0, cy_int - r_int - 1), min(h, cy_int + r_int + 2)
                x_min, x_max = max(0, cx_int - r_int - 1), min(w, cx_int + r_int + 2)
                
                Y, X = np.ogrid[y_min:y_max, x_min:x_max]
                dist_sq = (X - center_q[0])**2 + (Y - center_q[1])**2
                mask_circle = dist_sq <= radius**2
                
                roi_vals = brightness_image[y_min:y_max, x_min:x_max]
                # Handle edge cases where ROI might be smaller than mask due to image boundary
                mask_circle = mask_circle[:roi_vals.shape[0], :roi_vals.shape[1]]
                
                vals = roi_vals[mask_circle]
                
                if vals.size > 0:
                    means.append(round(float(np.mean(vals)), 2))
                    areas.append(vals.size)
                else:
                    means.append(0.0)
                    areas.append(0)

                # Visualization
                cv2.circle(vis_img, (cx_int, cy_int), r_int, color, 2)
        else:
            # No signal in particle
            means = [0.0] * 4
            areas = [0] * 4

        row_data = [pid] + means + areas
        results.append(row_data)

        # Draw axes
        p_cen = (int(centroid[0]), int(centroid[1]))
        p_y = (int(centroid[0] + ey[0] * 20), int(centroid[1] + ey[1] * 20))
        p_x = (int(centroid[0] + ex[0] * 20), int(centroid[1] + ex[1] * 20))
        cv2.line(vis_img, p_cen, p_y, (0, 0, 255), 2) # Red axis (Y -> Big Notch)
        cv2.line(vis_img, p_cen, p_x, (255, 0, 0), 2) # Blue axis (X -> Small Notch)
        cv2.putText(vis_img, str(pid), p_cen, cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

    # Calculate column means for the footer row
    mean_row = []
    sd_row = []
    sem_row = []
    if results:
        mean_row = ["Mean"]
        sd_row = ["SD"]
        sem_row = ["Std. Error"]
        # Columns 1 to 8 (Q1_Mean...Q4_Mean, Q1_Area...Q4_Area)
        for col_idx in range(1, 9):
            col_vals = [row[col_idx] for row in results]
            valid_vals = [v for v in col_vals if not np.isnan(v)]
            if valid_vals:
                mean_row.append(round(np.mean(valid_vals), 2))
                n = len(valid_vals)
                if n > 1:
                    sd = np.std(valid_vals, ddof=1)
                    sd_row.append(round(sd, 2))
                    sem_row.append(round(sd / np.sqrt(n), 2))
                else:
                    sd_row.append(0.0)
                    sem_row.append(0.0)
            else:
                mean_row.append(0.0)
                sd_row.append(0.0)
                sem_row.append(0.0)

    # Prepare footers
    footers = []
    if mean_row: footers.append(mean_row)
    if sd_row: footers.append(sd_row)
    if sem_row: footers.append(sem_row)

    header = ["particle_id", 
              "Q1_Mean", "Q2_Mean", "Q3_Mean", "Q4_Mean", 
              "Q1_Area", "Q2_Area", "Q3_Area", "Q4_Area"]

    # Write colored Excel
    _write_colored_xlsx(csv_path, header, results, footers)

    # Save population summary
    base, ext = os.path.splitext(csv_path)
    # _save_population_stats(results, f"{base}_summary{ext}")

    cv2.imwrite(overlay_path, vis_img)
    print(f"[OK] Masked measurement saved: {csv_path}")
    print(f"[OK] Visualization saved: {overlay_path}")


def compute_global_signal_intensity(
    brightness_image,
    structure_mask,
    signal_mask,
    csv_path="global_signal_results.csv",
    overlay_path="global_signal_visualization.png"
):
    """
    Calculates the mean intensity of signal pixels within each particle defined by structure_mask.
    No axes or quadrants are used.
    """
    h, w = brightness_image.shape

    # 1. Prepare structure labels
    num_labels, labels = cv2.connectedComponents(structure_mask)
    from skimage import measure
    regions = measure.regionprops(labels)
    # Sort by x-coordinate (col) for consistent IDs
    regions = sorted(regions, key=lambda r: r.centroid[1])
    region_id_map = {r.label: idx + 1 for idx, r in enumerate(regions)}

    # 2. Prepare visualization base
    min_val, max_val = brightness_image.min(), brightness_image.max()
    if max_val > min_val:
        display_base = 255.0 * (brightness_image - min_val) / (max_val - min_val)
    else:
        display_base = np.zeros_like(brightness_image)
    vis_img = cv2.cvtColor(display_base.astype(np.uint8), cv2.COLOR_GRAY2BGR)

    results = []

    for r in regions:
        label = r.label
        pid = region_id_map[label]
        
        minr, minc, maxr, maxc = r.bbox
        
        # Extract local masks to speed up
        local_structure = (labels[minr:maxr, minc:maxc] == label)
        local_signal = (signal_mask[minr:maxr, minc:maxc] > 0)
        local_brightness = brightness_image[minr:maxr, minc:maxc]
        
        # Intersection: pixels that are part of this particle AND are signal
        mask_intersection = local_structure & local_signal
        
        vals = local_brightness[mask_intersection]
        
        if vals.size > 0:
            mean_val = round(float(np.mean(vals)), 2)
            area_val = vals.size
        else:
            mean_val = 0.0
            area_val = 0
            
        results.append([pid, mean_val, area_val])

        # Visualization: Draw ID and Contour (Green)
        cy, cx = r.centroid
        cv2.putText(vis_img, str(pid), (int(cx), int(cy)), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        
        # Pad local_structure to ensure contours are closed and don't hit the boundary
        padded_structure = np.pad(local_structure, pad_width=1, mode='constant', constant_values=0)
        contours = measure.find_contours(padded_structure.astype(np.uint8), 0.5)
        for contour in contours:
            contour[:, 0] += (minr - 1)
            contour[:, 1] += (minc - 1)
            # cv2.polylines expects (x,y) points
            pts = np.expand_dims(np.flip(contour, axis=1).astype(np.int32), axis=1)
            cv2.drawContours(vis_img, [pts], -1, (0, 255, 0), 1)

    # Footer stats
    header = ["particle_id", "Mean_Signal_Intensity", "Signal_Area"]
    
    # Calculate stats for footer
    mean_row = ["Mean"]
    sd_row = ["SD"]
    sem_row = ["Std. Error"]

    if results:
        for col_idx in range(1, 3):
            col_vals = [row[col_idx] for row in results]
            valid_vals = [v for v in col_vals if not np.isnan(float(v))]
            
            if valid_vals:
                mean_row.append(round(np.mean(valid_vals), 2))
                n = len(valid_vals)
                if n > 1:
                    sd = np.std(valid_vals, ddof=1)
                    sd_row.append(round(sd, 2))
                    sem_row.append(round(sd / np.sqrt(n), 2))
                else:
                    sd_row.append(0.0)
                    sem_row.append(0.0)
            else:
                mean_row.append(0.0)
                sd_row.append(0.0)
                sem_row.append(0.0)
    else:
        mean_row.extend([0.0, 0.0])
        sd_row.extend([0.0, 0.0])
        sem_row.extend([0.0, 0.0])
    
    _write_colored_xlsx(csv_path, header, results, footer_rows=[mean_row, sd_row, sem_row])
    
    cv2.imwrite(overlay_path, vis_img)
    print(f"[OK] Global signal measurement saved: {csv_path}")
    print(f"[OK] Visualization saved: {overlay_path}")


def _save_population_stats(results, summary_path):
    """
    Calculate Mean, Std, and SEM across all particles (Inter-particle statistics)
    and save to a separate CSV.
    """
    # results structure: [pid, Q1_mean, Q2_mean, Q3_mean, Q4_mean, ...]
    # We extract indices 1, 2, 3, 4
    q_data = [[], [], [], []]

    for row in results:
        # row[0] is pid
        # row[1]..row[4] are Q1..Q4 means
        for i in range(4):
            val = row[1 + i]
            if not np.isnan(val):
                q_data[i].append(val)

    summary_rows = []
    # Headers
    headers = ["Statistic", "Q1", "Q2", "Q3", "Q4"]

    # Calculate stats
    means = ["Population_Mean"]
    stds = ["Population_Std"]
    sems = ["Population_SEM"]

    for i in range(4):
        arr = np.array(q_data[i])
        if arr.size > 0:
            m = np.mean(arr)
            s = np.std(arr, ddof=1) if arr.size > 1 else 0.0
            sem = s / np.sqrt(arr.size)

            means.append(round(m, 2))
            stds.append(round(s, 2))
            sems.append(round(sem, 2))
        else:
            means.append(np.nan)
            stds.append(np.nan)
            sems.append(np.nan)

    summary_rows = [means, stds, sems]

    with open(summary_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(summary_rows)

    print(f"[OK] Population statistics (Inter-particle) saved: {summary_path}")


def _write_colored_xlsx(path, headers, data_rows, footer_rows=None):
    """
    Writes data to an Excel file with colored columns based on Quadrant names.
    Q1: Blue, Q2: Green, Q3: Purple, Q4: Yellow
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("[WARNING] openpyxl not found. Saving as plain CSV without colors.")
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_rows)
            if footer_rows:
                writer.writerows(footer_rows)
        return

    # Ensure .xlsx extension
    if not path.lower().endswith(".xlsx"):
        path = os.path.splitext(path)[0] + ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Define colors (ARGB Hex)
    # Q1: Blue, Q2: Green, Q3: Purple, Q4: Dark Yellow (for readability on white)
    colors = {
        "Q1": "0000FF",
        "Q2": "008000",
        "Q3": "800080",
        "Q4": "CCCC00" 
    }

    def get_color(header_name):
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            if q in header_name:
                return colors[q]
        return None

    # Combine all rows: Header + Data + Footer
    all_rows = [headers] + data_rows + (footer_rows if footer_rows else [])

    for r_idx, row_data in enumerate(all_rows, 1):
        ws.append(row_data)
        for c_idx, val in enumerate(row_data):
            # Determine color based on the header of this column
            col_header = headers[c_idx]
            c_code = get_color(col_header)
            if c_code:
                cell = ws.cell(row=r_idx, column=c_idx+1)
                # Bold for header (row 1), regular for others
                cell.font = Font(color=c_code, bold=(r_idx==1))
            elif r_idx == 1:
                # Bold for non-colored headers (e.g. particle_id)
                ws.cell(row=r_idx, column=c_idx+1).font = Font(bold=True)

    wb.save(path)
    print(f"[OK] Colored Excel table saved: {path}")
