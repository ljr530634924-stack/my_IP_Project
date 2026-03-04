import cv2
import numpy as np
import os
import glob
import sys
import csv
from scipy import ndimage
from scipy.spatial import distance as sp_dist
from skimage import morphology

# 尝试导入 openpyxl 以生成 Excel 文件，如果不存在则回退到 CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not found. Output will be CSV instead of XLSX.")

# 导入现有的图像增强函数
from handle_ch01 import adjust_ch01_image

# --- Configuration ---
# 输入文件夹路径
INPUT_FOLDER = r"D:\Ingenieurpraixs\test_p3"

# ROI 参数
ROI_DIAMETER_RATIO = 0.85  # ROI 直径占图像短边的比例 (0.85 = 85%)

SAVE_DEBUG_IMAGES = False   # [New] 是否保存调试图片

# 信号检测参数 (基于增强后的图像)
THRESHOLD_RATIO = 0.6      # [Modified] 恢复为 0.6 以匹配 run_NN_DB_SM
MIN_AREA = 1000            # [Modified] 降低为 10 以匹配 run_NN_DB_SM 的 SIGNAL_MIN_AREA
MAX_AREA = 12000           # [Modified] 增加最大面积以匹配 find_MC 的默认值
MIN_CIRCULARITY = 0.4     # [Modified] 降低圆度要求 (0.5 -> 0.1) 防止小点被过滤
SMOOTHING_RADIUS = 3       # [New] 平滑半径 (建议 2-4)。用于消除边缘锯齿，提高圆度计算准确性。

# 测量参数 (基于原始图像)
MEASURE_RADIUS = 48       # 固定测量圆半径 (Area ~ 8171)

# [New] Post-processing filters for detected signals
REMOVE_OVERLAPPING = True           # 如果为 True, 则删除相互重叠的圆
ISOLATION_THRESHOLD_RATIO = 2.5    # 删除在此距离（测量半径的倍数）内没有邻居的圆。设为 0 可禁用。

# 可视化参数
FONT_SCALE = 0.8

def get_roi_mask(shape, ratio):
    """生成中心圆形的 ROI Mask"""
    h, w = shape[:2]
    mask = np.zeros((h, w), dtype=np.uint8)
    center = (w // 2, h // 2)
    radius = int(min(h, w) * ratio / 2)
    cv2.circle(mask, center, radius, 255, -1)
    return mask, center, radius

def calculate_circularity(area, perimeter):
    if perimeter == 0: return 0
    return 4 * np.pi * area / (perimeter * perimeter)


def process_image(ch01_path):
    directory = os.path.dirname(ch01_path)
    filename = os.path.basename(ch01_path)
    name_no_ext = os.path.splitext(filename)[0]
    
    print(f"Processing: {filename}")

    # 1. 图像增强 (用于检测)
    # 使用临时文件保存增强后的图像
    temp_adj_path = os.path.join(directory, f"temp_adj_{name_no_ext}.png")
    
    # [Modified] 参数设置：与 run_NN_DB_SM.py 保持一致
    adjust_ch01_image(
        ch01_path, 
        temp_adj_path,
        exposure=1.5,       
        brightness=10,
        contrast_gain=1.0,
        stretch_low=20.0,
        stretch_high=98.5,
        do_stretch=True
    )
    
    adj_img = cv2.imread(temp_adj_path, cv2.IMREAD_GRAYSCALE)
    if adj_img is None:
        print(f"  [ERROR] Failed to read adjusted image: {temp_adj_path}")
        return None, None

    if SAVE_DEBUG_IMAGES:
        cv2.imwrite(os.path.join(directory, f"{name_no_ext}_debug_01_enhanced.png"), adj_img)

    # 2. 创建 ROI Mask
    roi_mask, roi_center, roi_radius = get_roi_mask(adj_img.shape, ROI_DIAMETER_RATIO)
    if SAVE_DEBUG_IMAGES:
        cv2.imwrite(os.path.join(directory, f"{name_no_ext}_debug_02_roi_mask.png"), roi_mask)
    
    # 3. 阈值处理
    # [Modified] 使用比例计算阈值
    threshold_val = int(THRESHOLD_RATIO * 255)
    _, binary = cv2.threshold(adj_img, threshold_val, 255, cv2.THRESH_BINARY)
    
    # [New] 匹配 run_NN_DB_SM 的形态学处理逻辑
    # 3.1 去除小噪点
    binary_bool = morphology.remove_small_objects(binary > 0, min_size=MIN_AREA)
    binary_cleaned = (binary_bool.astype(np.uint8) * 255)
    
    # 3.2 填充孔洞
    binary_filled = (ndimage.binary_fill_holes(binary_cleaned > 0) * 255).astype(np.uint8)

    # [New] 边缘平滑处理：消除锯齿状边缘 (Coastline effect)
    if SMOOTHING_RADIUS > 0:
        # 使用椭圆核进行开运算 (Opening)，可以磨平边缘毛刺
        k_size = SMOOTHING_RADIUS * 2 + 1
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (k_size, k_size))
        binary_filled = cv2.morphologyEx(binary_filled, cv2.MORPH_OPEN, kernel)

    if SAVE_DEBUG_IMAGES:
        cv2.imwrite(os.path.join(directory, f"{name_no_ext}_debug_03_signal_mask_filled.png"), binary_filled)
    
    # 4. 应用 ROI (只保留中心区域的信号)
    binary_roi = cv2.bitwise_and(binary_filled, binary_filled, mask=roi_mask)
    
    # 5. 查找轮廓
    contours, _ = cv2.findContours(binary_roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print(f"  [DEBUG] Found {len(contours)} contours before filtering.")
    
    # 6. 筛选与测量
    raw_img = cv2.imread(ch01_path, cv2.IMREAD_UNCHANGED) # 读取原始数据用于测量
    
    # 创建可视化图 (RGB) - Ch01
    vis_img_01 = cv2.cvtColor(adj_img, cv2.COLOR_GRAY2BGR)
    # 画出 ROI 边界 (青色)
    cv2.circle(vis_img_01, roi_center, roi_radius, (255, 255, 0), 2)
    
    # [New] 尝试读取 Ch00 用于可视化
    ch00_path = ch01_path.replace("ch01", "ch00")
    vis_img_00 = None
    if os.path.exists(ch00_path):
        ch00_raw = cv2.imread(ch00_path, cv2.IMREAD_UNCHANGED)
        if ch00_raw is not None:
            # 归一化以便显示
            if ch00_raw.dtype == np.uint16:
                vis_img_00 = (ch00_raw / 256).astype(np.uint8)
            else:
                vis_img_00 = ch00_raw.astype(np.uint8)
            vis_img_00 = cv2.cvtColor(vis_img_00, cv2.COLOR_GRAY2BGR)
            # 在 Ch00 上也画出 ROI
            cv2.circle(vis_img_00, roi_center, roi_radius, (255, 255, 0), 2)
    
    measurements = []
    
    # [Modified] 直接使用原始轮廓，不进行分水岭分割
    processed_contours = contours
            
    if SAVE_DEBUG_IMAGES:
        # 保存分割后的 Mask 供检查
        debug_filtered = np.zeros_like(binary_roi)
        cv2.drawContours(debug_filtered, processed_contours, -1, 255, -1)
        cv2.imwrite(os.path.join(directory, f"{name_no_ext}_debug_04_signal_mask_filtered.png"), debug_filtered)

    # --- [Modified] Step 1: Initial filtering and collection ---
    # This loop now only collects potential candidates. Drawing is done after post-processing.
    for i, cnt in enumerate(processed_contours):
        area = cv2.contourArea(cnt)
        
        # 过滤极小噪点
        if area < MIN_AREA: continue
        
        # 过滤过大粘连点 (用红色标记被丢弃的粘连点)
        if area > MAX_AREA:
            cv2.drawContours(vis_img_01, [cnt], -1, (0, 0, 255), 1)
            continue
            
        perimeter = cv2.arcLength(cnt, True)
        circularity = calculate_circularity(area, perimeter)
        
        # 过滤形状不规则的点 (用橙色标记)
        if circularity < MIN_CIRCULARITY:
            cv2.drawContours(vis_img_01, [cnt], -1, (0, 165, 255), 1)
            continue
            
        M = cv2.moments(cnt)
        if M["m00"] == 0: continue
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        
        # 在原始图像上测量固定圆内的强度
        mask_c = np.zeros(raw_img.shape[:2], dtype=np.uint8)
        cv2.circle(mask_c, (cx, cy), MEASURE_RADIUS, 255, -1)
        mean_val = cv2.mean(raw_img, mask=mask_c)[0]
        
        measurements.append({
            "cx": cx,
            "cy": cy,
            "area": area,
            "circularity": circularity,
            "mean_intensity": mean_val
        })

    # --- [New] Step 2: Post-processing filters (Overlap and Isolation) ---
    print(f"  -> Found {len(measurements)} candidate signals before post-processing.")
    final_measurements = measurements

    # Filter 1: Remove overlapping circles (run first)
    if REMOVE_OVERLAPPING and len(final_measurements) > 1:
        points = np.array([[m['cx'], m['cy']] for m in final_measurements])
        dist_matrix = sp_dist.squareform(sp_dist.pdist(points))
        np.fill_diagonal(dist_matrix, np.inf)
        
        overlap_threshold = 2 * MEASURE_RADIUS
        overlapping_pairs = np.argwhere(dist_matrix < overlap_threshold)
        indices_to_remove = np.unique(overlapping_pairs.flatten())
        
        if len(indices_to_remove) > 0:
            keep_mask = np.ones(len(final_measurements), dtype=bool)
            keep_mask[indices_to_remove] = False
            final_measurements = [m for i, m in enumerate(final_measurements) if keep_mask[i]]
            print(f"  -> Removed {len(indices_to_remove)} overlapping signals.")

    # Filter 2: Remove isolated circles (run second on the remaining points)
    if ISOLATION_THRESHOLD_RATIO > 0 and len(final_measurements) > 1:
        points = np.array([[m['cx'], m['cy']] for m in final_measurements])
        dist_matrix = sp_dist.squareform(sp_dist.pdist(points))
        np.fill_diagonal(dist_matrix, np.inf)
        
        min_distances = np.min(dist_matrix, axis=1)
        isolation_threshold = ISOLATION_THRESHOLD_RATIO * MEASURE_RADIUS
        is_isolated_mask = min_distances > isolation_threshold
        
        num_isolated = np.sum(is_isolated_mask)
        if num_isolated > 0:
            final_measurements = [m for i, m in enumerate(final_measurements) if not is_isolated_mask[i]]
            print(f"  -> Removed {num_isolated} isolated signals.")

    # --- [New] Step 3: Finalize IDs and draw visualizations ---
    for i, m in enumerate(final_measurements):
        m['id'] = i + 1
        cx, cy = m['cx'], m['cy']
        # 可视化：画绿色测量圆 (Ch01)
        cv2.circle(vis_img_01, (cx, cy), MEASURE_RADIUS, (0, 255, 0), 2)
        cv2.putText(vis_img_01, str(m['id']), (cx, cy), cv2.FONT_HERSHEY_SIMPLEX, FONT_SCALE, (0, 255, 0), 2)
        # 可视化：画绿色测量圆 (Ch00)
        if vis_img_00 is not None:
            cv2.circle(vis_img_00, (cx, cy), MEASURE_RADIUS, (0, 255, 0), 2)
            cv2.putText(vis_img_00, str(m['id']), (cx, cy), cv2.FONT_HERSHEY_SIMPLEX, FONT_SCALE, (0, 255, 0), 2)

    # 清理临时文件
    if os.path.exists(temp_adj_path):
        os.remove(temp_adj_path)
        
    # 保存可视化图 (Ch01)
    vis_path_01 = None
    if SAVE_DEBUG_IMAGES:
        vis_path_01 = os.path.join(directory, f"{name_no_ext}_direct_vis_01.png")
        cv2.imwrite(vis_path_01, vis_img_01)
    
    # 保存可视化图 (Ch00)
    vis_path_00 = None
    if vis_img_00 is not None:
        vis_name_00 = name_no_ext.replace("_ch01", "")
        vis_path_00 = os.path.join(directory, f"{vis_name_00}_00visualization.png")
        cv2.imwrite(vis_path_00, vis_img_00)
    
    return final_measurements, vis_path_01, vis_path_00

def save_results(measurements, output_path):
    if not measurements:
        return

    # 计算全局平均值
    all_means = [m["mean_intensity"] for m in measurements]
    global_avg = np.mean(all_means) if all_means else 0
    global_std = np.std(all_means) if all_means else 0
    
    headers = ["Signal_ID", "Centroid_X", "Centroid_Y", "Contour_Area", "Circularity", "Mean_Intensity"]
    
    if HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = "Signal Data"
        
        ws.append(headers)
        # 设置表头加粗
        for cell in ws[1]: cell.font = Font(bold=True)
            
        for m in measurements:
            ws.append([m["id"], m["cx"], m["cy"], m["area"], m["circularity"], m["mean_intensity"]])
            
        # 添加统计摘要
        ws.append([])
        ws.append(["Summary Statistics"])
        ws.append(["Total Signals", len(measurements)])
        ws.append(["Global Average Intensity", global_avg])
        ws.append(["Std Dev", global_std])
        
        # 加粗统计部分
        for i in range(4):
            ws[f"A{ws.max_row - i}"].font = Font(bold=True)
        
        wb.save(output_path)
    else:
        # CSV 回退
        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for m in measurements:
                writer.writerow([m["id"], m["cx"], m["cy"], m["area"], m["circularity"], m["mean_intensity"]])
            writer.writerow([])
            writer.writerow(["Summary Statistics"])
            writer.writerow(["Global Average Intensity", global_avg])

def main():
    print(f"=== Starting Direct Signal Measurement in '{INPUT_FOLDER}' ===")
    
    # 直接查找 ch01 文件
    search_pattern = os.path.join(INPUT_FOLDER, "*ch01*.tif")
    files = glob.glob(search_pattern)
    
    if not files:
        print("No ch01 files found.")
        return
        
    print(f"Found {len(files)} files.")
    
    for f in files:
        results, vis_path_01, vis_path_00 = process_image(f)
        
        if results:
            base_name = os.path.splitext(os.path.basename(f))[0]
            xlsx_path = os.path.join(os.path.dirname(f), f"{base_name}_direct_results.xlsx")
                
            save_results(results, xlsx_path)
            
            print(f"  -> Saved results: {os.path.basename(xlsx_path)}")
            if vis_path_01:
                print(f"  -> Saved visualization 01: {os.path.basename(vis_path_01)}")
            if vis_path_00:
                print(f"  -> Saved visualization 00: {os.path.basename(vis_path_00)}")
            print(f"  -> Stats: Count={len(results)}, Avg Intensity={np.mean([m['mean_intensity'] for m in results]):.2f}")
        else:
            print("  -> No valid signals found.")

if __name__ == "__main__":
    main()