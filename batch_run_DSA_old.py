import cv2
import numpy as np
import os
import glob
import sys
import csv
from scipy import ndimage
from scipy.spatial import distance as sp_dist
from skimage import morphology

# Try importing openpyxl for Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not found. Output will be CSV instead of XLSX.")

# Import existing image adjustment function
from handle_ch01 import adjust_ch01_image

# --- Configuration ---
# Set this to the folder containing your images.
INPUT_FOLDER = r"D:\Ingenieurpraixs\test_p3"

# ROI Parameters
ROI_DIAMETER_RATIO = 0.85  # ROI diameter ratio relative to the shorter side of the image

# Signal Detection Parameters (Based on enhanced ch01 image)
THRESHOLD_RATIO = 0.6      # Threshold ratio (0.0-1.0)
MIN_AREA = 1000            # Minimum area to keep a signal
MAX_AREA = 15000           # Maximum area to keep a signal
MIN_CIRCULARITY = 0.4      # Minimum circularity (0.0 - 1.0)
SMOOTHING_RADIUS = 3       # Smoothing radius for morphological opening

# Measurement Parameters (Based on raw ch01 image)
MEASURE_RADIUS = 70       # Fixed measurement circle radius

# Post-processing Filters
REMOVE_OVERLAPPING = True           # Remove overlapping circles
ISOLATION_THRESHOLD_RATIO = 3    # Remove circles without neighbors within this distance ratio
# Visualization
FONT_SCALE = 0.8


def get_roi_mask(shape, ratio):
    """Generate centered circular ROI Mask"""
    h, w = shape[:2]
    mask = np.zeros((h, w), dtype=np.uint8)
    center = (w // 2, h // 2)
    radius = int(min(h, w) * ratio / 2)
    cv2.circle(mask, center, radius, 255, -1)
    return mask, center, radius

def calculate_circularity(area, perimeter):
    if perimeter == 0: return 0
    return 4 * np.pi * area / (perimeter * perimeter)

def save_results(measurements, output_path):
    if not measurements:
        return

    all_means = [m["mean_intensity"] for m in measurements]
    global_avg = np.mean(all_means) if all_means else 0
    global_std = np.std(all_means) if all_means else 0
    
    headers = ["Signal_ID", "Centroid_X", "Centroid_Y", "Contour_Area", "Circularity", "Mean_Intensity"]
    
    if HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = "Signal Data"
        
        ws.append(headers)
        for cell in ws[1]: cell.font = Font(bold=True)
            
        for m in measurements:
            ws.append([m["id"], m["cx"], m["cy"], m["area"], m["circularity"], m["mean_intensity"]])
            
        ws.append([])
        ws.append(["Summary Statistics"])
        ws.append(["Total Signals", len(measurements)])
        ws.append(["Global Average Intensity", global_avg])
        ws.append(["Std Dev", global_std])
        
        for i in range(4):
            ws[f"A{ws.max_row - i}"].font = Font(bold=True)
        
        wb.save(output_path)
    else:
        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for m in measurements:
                writer.writerow([m["id"], m["cx"], m["cy"], m["area"], m["circularity"], m["mean_intensity"]])
            writer.writerow([])
            writer.writerow(["Summary Statistics"])
            writer.writerow(["Global Average Intensity", global_avg])

def process_pair(ch00_path, ch01_path):
    print(f"Processing pair:\n  CH00: {os.path.basename(ch00_path)}\n  CH01: {os.path.basename(ch01_path)}")
    
    directory = os.path.dirname(ch00_path)
    base_name = os.path.basename(ch00_path)
    name_no_ext = os.path.splitext(base_name)[0]
    
    # Create prefix
    prefix = name_no_ext.replace("ch00", "")
    prefix = prefix.replace("__", "_").strip(" _")
    if not prefix:
        prefix = "output"
        
    # Output paths
    csv_output = os.path.join(directory, f"{prefix}_DSA_results.xlsx")
    vis_output = os.path.join(directory, f"{prefix}_DSA_visualization.png")
    
    # --- 1. Prepare Visualization Background (Ch00) ---
    # Ch00 is used ONLY for visualization background
    ch00_raw = cv2.imread(ch00_path, cv2.IMREAD_UNCHANGED)
    if ch00_raw is None:
        print(f"  [ERROR] Cannot read {ch00_path}")
        return

    # Normalize Ch00 for display (handle 16-bit)
    if ch00_raw.dtype == np.uint16:
        vis_img = (ch00_raw / 256).astype(np.uint8)
    else:
        vis_img = ch00_raw.astype(np.uint8)
    
    if vis_img.ndim == 2:
        vis_img = cv2.cvtColor(vis_img, cv2.COLOR_GRAY2BGR)

    # --- 2. Process Ch01 for Signal Detection ---
    # Enhance Ch01
    temp_adj_path = os.path.join(directory, f"temp_adj_{prefix}.png")
    adjust_ch01_image(
        ch01_path, 
        temp_adj_path,
        exposure=3.0,       
        brightness=10,
        contrast_gain=1.0,
        stretch_low=2,
        stretch_high=98.5,
        do_stretch=True
    )
    
    adj_img = cv2.imread(temp_adj_path, cv2.IMREAD_GRAYSCALE)
    if adj_img is None:
        print(f"  [ERROR] Failed to read adjusted image: {temp_adj_path}")
        return

    # Create ROI Mask
    roi_mask, roi_center, roi_radius = get_roi_mask(adj_img.shape, ROI_DIAMETER_RATIO)
    
    # Draw ROI on visualization (Cyan)
    cv2.circle(vis_img, roi_center, roi_radius, (255, 255, 0), 2)

    # Thresholding
    threshold_val = int(THRESHOLD_RATIO * 255)
    _, binary = cv2.threshold(adj_img, threshold_val, 255, cv2.THRESH_BINARY)
    
    # Morphology
    binary_bool = morphology.remove_small_objects(binary > 0, min_size=MIN_AREA)
    binary_cleaned = (binary_bool.astype(np.uint8) * 255)
    binary_filled = (ndimage.binary_fill_holes(binary_cleaned > 0) * 255).astype(np.uint8)

    if SMOOTHING_RADIUS > 0:
        k_size = SMOOTHING_RADIUS * 2 + 1
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (k_size, k_size))
        binary_filled = cv2.morphologyEx(binary_filled, cv2.MORPH_OPEN, kernel)

    # Apply ROI
    binary_roi = cv2.bitwise_and(binary_filled, binary_filled, mask=roi_mask)
    
    # Find Contours
    contours, _ = cv2.findContours(binary_roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # --- 3. Filter and Measure ---
    raw_ch01 = cv2.imread(ch01_path, cv2.IMREAD_UNCHANGED)
    if raw_ch01 is None:
        print(f"  [ERROR] Cannot read raw {ch01_path}")
        if os.path.exists(temp_adj_path): os.remove(temp_adj_path)
        return

    measurements = []
    
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < MIN_AREA: continue
        if area > MAX_AREA: continue
            
        perimeter = cv2.arcLength(cnt, True)
        circularity = calculate_circularity(area, perimeter)
        
        if circularity < MIN_CIRCULARITY: continue
            
        M = cv2.moments(cnt)
        if M["m00"] == 0: continue
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        
        # Measure on raw Ch01
        mask_c = np.zeros(raw_ch01.shape[:2], dtype=np.uint8)
        cv2.circle(mask_c, (cx, cy), MEASURE_RADIUS, 255, -1)
        mean_val = cv2.mean(raw_ch01, mask=mask_c)[0]
        
        measurements.append({
            "cx": cx,
            "cy": cy,
            "area": area,
            "circularity": circularity,
            "mean_intensity": mean_val
        })

    # --- 4. Post-processing ---
    final_measurements = measurements

    # Filter 1: Overlap
    if REMOVE_OVERLAPPING and len(final_measurements) > 1:
        points = np.array([[m['cx'], m['cy']] for m in final_measurements])
        dist_matrix = sp_dist.squareform(sp_dist.pdist(points))
        np.fill_diagonal(dist_matrix, np.inf)
        overlap_threshold = 2 * MEASURE_RADIUS
        overlapping_pairs = np.argwhere(dist_matrix < overlap_threshold)
        indices_to_remove = np.unique(overlapping_pairs.flatten())
        if len(indices_to_remove) > 0:
            keep_mask = np.ones(len(final_measurements), dtype=bool)
            keep_mask[indices_to_remove] = False
            final_measurements = [m for i, m in enumerate(final_measurements) if keep_mask[i]]

    # Filter 2: Isolation
    if ISOLATION_THRESHOLD_RATIO > 0 and len(final_measurements) > 1:
        points = np.array([[m['cx'], m['cy']] for m in final_measurements])
        dist_matrix = sp_dist.squareform(sp_dist.pdist(points))
        np.fill_diagonal(dist_matrix, np.inf)
        min_distances = np.min(dist_matrix, axis=1)
        isolation_threshold = ISOLATION_THRESHOLD_RATIO * MEASURE_RADIUS
        is_isolated_mask = min_distances > isolation_threshold
        if np.sum(is_isolated_mask) > 0:
            final_measurements = [m for i, m in enumerate(final_measurements) if not is_isolated_mask[i]]

    # --- 5. Draw and Save ---
    for i, m in enumerate(final_measurements):
        m['id'] = i + 1
        cx, cy = m['cx'], m['cy']
        # Draw on Ch00 visualization (Green)
        cv2.circle(vis_img, (cx, cy), MEASURE_RADIUS, (0, 255, 0), 2)
        cv2.putText(vis_img, str(m['id']), (cx, cy), cv2.FONT_HERSHEY_SIMPLEX, FONT_SCALE, (0, 255, 0), 2)

    save_results(final_measurements, csv_output)
    cv2.imwrite(vis_output, vis_img)
    
    print(f"  -> Saved results: {os.path.basename(csv_output)}")
    print(f"  -> Saved visualization: {os.path.basename(vis_output)}")
    print(f"  -> Signals found: {len(final_measurements)}")

    # Cleanup
    if os.path.exists(temp_adj_path):
        os.remove(temp_adj_path)

def main():
    print(f"=== Starting Batch DSA Analysis in '{os.path.abspath(INPUT_FOLDER)}' ===")
    
    search_pattern = os.path.join(INPUT_FOLDER, "*ch00*.tif")
    ch00_files = glob.glob(search_pattern)
    
    if not ch00_files:
        print(f"No files found matching {search_pattern}")
        return

    print(f"Found {len(ch00_files)} candidate ch00 files.")

    count = 0
    for ch00 in ch00_files:
        directory, filename = os.path.split(ch00)
        filename_ch01 = filename.replace("ch00", "ch01")
        ch01 = os.path.join(directory, filename_ch01)
        
        if not os.path.exists(ch01):
            print(f"[WARN] Corresponding ch01 file not found for {filename}. Skipping.")
            continue
            
        process_pair(ch00, ch01)
        count += 1

    print(f"\n=== Batch Processing Complete! Processed {count} pairs. ===")

if __name__ == "__main__":
    main()